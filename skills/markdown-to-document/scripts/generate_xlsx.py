"""
Generate an Excel workbook from JSON input.

Usage:
    python generate_xlsx.py input.json -o output.xlsx
    python generate_xlsx.py input.json -t template.xlsx -o output.xlsx
    python generate_xlsx.py input.json -t template.xlsx -o output.xlsx --mode inject

Modes:
    generate (default) — creates/fills workbook, may overwrite cell formatting
    inject             — preserves ALL cell formatting (font, fill, border,
                         number_format, alignment, protection) while replacing values

Input format (JSON) — generate mode:
    {
      "sheets": {
        "SheetName": {
          "headers": ["Col1", "Col2", "Col3"],
          "rows": [
            ["val1", "val2", "val3"],
            ["val4", "val5", "val6"]
          ]
        }
      }
    }

Input format (JSON) — inject mode:
    {
      "sheets": {
        "SheetName": {
          "start_row": 2,
          "start_col": 1,
          "headers": ["Col1", "Col2", "Col3"],
          "rows": [
            ["val1", "val2", "val3"],
            ["val4", "val5", "val6"]
          ]
        }
      }
    }

Requires: pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
from copy import copy
from pathlib import Path


def _ensure_package(import_name: str, pip_name: str | None = None) -> bool:
    """Try to import a package; auto-install if non-interactive. Returns True if available."""
    pip_name = pip_name or import_name
    try:
        __import__(import_name)
        return True
    except (ImportError, OSError):
        pass

    if sys.stdin.isatty():
        answer = input(f'Package "{pip_name}" is not installed. Install it now? [Y/n] ').strip().lower()
        if answer in ("", "y", "yes"):
            subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name])
            return True
        return False
    else:
        print(f"Auto-installing missing package: {pip_name}", file=sys.stderr)
        subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name])
        __import__(import_name)  # verify
        return True


def load_input(input_path: Path) -> dict:
    """Load and validate JSON input."""
    text = input_path.read_text(encoding="utf-8")
    data = json.loads(text)
    if "sheets" not in data:
        print("Error: JSON must have a 'sheets' key.", file=sys.stderr)
        sys.exit(1)
    return data["sheets"]


def generate_from_scratch(sheets: dict, output_path: Path) -> None:
    """Create a new workbook from JSON data."""
    if not _ensure_package("openpyxl"):
        print("Error: openpyxl is required.", file=sys.stderr)
        sys.exit(1)
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    first_sheet = True

    for sheet_name, sheet_data in sheets.items():
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_name)

        headers = sheet_data.get("headers", [])
        rows = sheet_data.get("rows", [])

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                parsed = _parse_value(value)
                ws.cell(row=row_idx, column=col_idx, value=parsed)

        # Auto-width columns
        for col_idx in range(1, len(headers) + 1):
            max_len = 0
            for row_idx in range(1, len(rows) + 2):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    sheet_count = len(sheets)
    row_count = sum(len(s.get("rows", [])) for s in sheets.values())
    print(f"Generated: {output_path} ({sheet_count} sheets, {row_count} data rows)")


def generate_with_template(sheets: dict, template_path: Path, output_path: Path) -> None:
    """Fill an existing template workbook with JSON data."""
    if not _ensure_package("openpyxl"):
        print("Error: openpyxl is required.", file=sys.stderr)
        sys.exit(1)
    from openpyxl import load_workbook

    wb = load_workbook(str(template_path))

    for sheet_name, sheet_data in sheets.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        headers = sheet_data.get("headers", [])
        rows = sheet_data.get("rows", [])

        # Write headers in row 1
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Write data starting from row 2
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                parsed = _parse_value(value)
                ws.cell(row=row_idx, column=col_idx, value=parsed)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Generated: {output_path} (template: {template_path.name})")


def _inject_cell_value(ws, row: int, col: int, value) -> None:
    """Set cell value while preserving all formatting. Skips MergedCell objects."""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return  # Cannot write to non-origin merged cells
    # Capture existing formatting before setting value
    old_font = copy(cell.font)
    old_fill = copy(cell.fill)
    old_border = copy(cell.border)
    old_number_format = cell.number_format
    old_alignment = copy(cell.alignment)
    old_protection = copy(cell.protection)

    # Set new value
    cell.value = value

    # Restore formatting
    cell.font = old_font
    cell.fill = old_fill
    cell.border = old_border
    cell.number_format = old_number_format
    cell.alignment = old_alignment
    cell.protection = old_protection


def inject_with_template(sheets: dict, template_path: Path, output_path: Path) -> None:
    """Inject data into a template workbook preserving ALL cell formatting."""
    if not _ensure_package("openpyxl"):
        print("Error: openpyxl is required.", file=sys.stderr)
        sys.exit(1)
    from openpyxl import load_workbook

    wb = load_workbook(str(template_path))
    total_injected = 0

    for sheet_name, sheet_data in sheets.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # Sheet doesn't exist in template — create and fill normally
            ws = wb.create_sheet(sheet_name)
            headers = sheet_data.get("headers", [])
            rows = sheet_data.get("rows", [])
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=_parse_value(value))
                    total_injected += 1
            continue

        start_row = sheet_data.get("start_row", 2)
        start_col = sheet_data.get("start_col", 1)
        headers = sheet_data.get("headers", [])
        rows = sheet_data.get("rows", [])

        # Inject headers (in the row before start_row, or row 1 if start_row=2)
        if headers:
            header_row = start_row - 1 if start_row > 1 else 1
            for col_offset, header in enumerate(headers):
                col = start_col + col_offset
                _inject_cell_value(ws, header_row, col, header)
                total_injected += 1

        # Inject data rows
        for row_offset, row_data in enumerate(rows):
            row = start_row + row_offset
            for col_offset, value in enumerate(row_data):
                col = start_col + col_offset
                parsed = _parse_value(value)
                _inject_cell_value(ws, row, col, parsed)
                total_injected += 1

        # Clear leftover rows (preserve formatting, set value=None)
        data_end_row = start_row + len(rows) - 1
        max_col = start_col + max(len(r) for r in rows) - 1 if rows else start_col
        # Scan downward from data_end_row+1 to find rows with values
        check_row = data_end_row + 1
        while check_row <= ws.max_row:
            has_value = False
            for c in range(start_col, max_col + 1):
                if ws.cell(row=check_row, column=c).value is not None:
                    has_value = True
                    break
            if not has_value:
                break
            # Clear values but preserve formatting
            for c in range(start_col, max_col + 1):
                _inject_cell_value(ws, check_row, c, None)
            check_row += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    sheet_count = len(sheets)
    print(f"Injected: {output_path} ({sheet_count} sheets, {total_injected} cells, template: {template_path.name})")


def _parse_value(value):
    """Attempt to convert string values to appropriate Python types."""
    if not isinstance(value, str):
        return value
    # Try int
    try:
        return int(value)
    except (ValueError, TypeError):
        pass
    # Try float
    try:
        return float(value)
    except (ValueError, TypeError):
        pass
    return value


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate an Excel workbook from JSON input."
    )
    parser.add_argument(
        "input",
        type=Path,
        help="Path to JSON input file."
    )
    parser.add_argument(
        "-t", "--template",
        type=Path,
        default=None,
        help="Path to an .xlsx template file. Data is injected into matching sheets."
    )
    parser.add_argument(
        "-o", "--output",
        type=Path,
        default=Path("output.xlsx"),
        help="Output file path (default: output.xlsx)."
    )
    parser.add_argument(
        "--mode",
        choices=["generate", "inject"],
        default="generate",
        help="Mode: 'generate' (default) fills template/creates workbook; "
             "'inject' preserves all cell formatting while replacing values."
    )
    args = parser.parse_args()

    if not args.input.is_file():
        print(f"Error: Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)
    if args.template and not args.template.is_file():
        print(f"Error: Template file not found: {args.template}", file=sys.stderr)
        sys.exit(1)
    if args.mode == "inject" and not args.template:
        print("Error: --mode inject requires a template (-t).", file=sys.stderr)
        sys.exit(1)

    sheets = load_input(args.input)

    if args.mode == "inject":
        inject_with_template(sheets, args.template, args.output)
    elif args.template:
        generate_with_template(sheets, args.template, args.output)
    else:
        generate_from_scratch(sheets, args.output)


if __name__ == "__main__":
    main()
